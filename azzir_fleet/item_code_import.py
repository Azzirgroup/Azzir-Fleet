# Copyright (c) 2026, Azzir and contributors
# For license information, please see license.txt
"""Bulk-load Item alternative part numbers into the azzir_alias_codes child table
(Item Code Entry) using batched SQL — built for tens of thousands of rows.

For each Item in the spreadsheet it REPLACES that item's code table with the
sheet's rows: the item's own code is kept as the single Primary (so the item is
never renamed) and the alternatives are added. Codes already owned by a DIFFERENT
item are skipped and reported (the app forbids reusing a code).

Spreadsheet columns (row 1 = header):
    ID | azzir_alias_codes.code | azzir_alias_codes.is_primary | azzir_alias_codes.source

Used by both the CLI script (import_item_codes.py) and the "Item Code Import"
DocType (upload + run from the desk, incl. on Frappe Cloud)."""

import frappe
from frappe.utils import now_datetime

CHILD_DT = "Item Code Entry"
PARENT_FIELD = "azzir_alias_codes"
DELETE_BATCH = 500
INSERT_CHUNK = 5000
DEFAULT_SHEET = "Item Codes"


def parse_workbook(path_or_stream, sheet=None):
	"""{item_id: [(code, is_primary, source), ...]} preserving row order + the id order."""
	import openpyxl

	wb = openpyxl.load_workbook(path_or_stream, read_only=True, data_only=True)
	name = sheet or DEFAULT_SHEET
	ws = wb[name] if name in wb.sheetnames else wb.worksheets[0]
	rows = ws.iter_rows(min_row=1, values_only=True)
	next(rows, None)  # header
	groups, order = {}, []
	for r in rows:
		if not r or r[0] in (None, ""):
			continue
		item = str(r[0]).strip()
		code = str(r[1]).strip() if len(r) > 1 and r[1] is not None else ""
		if not code:
			continue
		is_primary = 1 if (len(r) > 2 and r[2] in (1, "1", True)) else 0
		source = (r[3] if len(r) > 3 and r[3] else "Manual")
		if item not in groups:
			groups[item] = []
			order.append(item)
		groups[item].append((code, is_primary, source))
	return groups, order


def build_plan(groups, order):
	"""Resolve conflicts. Returns (import_items, desired, stats).
	desired = {item: [(code, is_primary, source, idx), ...]}."""
	item_names = set(frappe.db.sql_list("select name from `tabItem`"))
	import_items = [i for i in order if i in item_names]
	missing = [i for i in order if i not in item_names]
	import_set = set(import_items)

	# Codes reserved by items NOT being reimported: their own name + their aliases.
	reserved = {}
	for nm in item_names:
		if nm not in import_set:
			reserved[nm] = nm
	for code, parent in frappe.db.sql(
		"select code, parent from `tabItem Code Entry` where parenttype = 'Item'"
	):
		if parent not in import_set:
			reserved.setdefault(code, parent)

	assigned = {}
	desired = {}
	conflicts = []
	for item in import_items:
		out, seen, idx = [], set(), 0
		for code, is_primary, source in groups[item]:
			if code in seen:
				continue
			if code in reserved and reserved[code] != item:
				conflicts.append((item, code, "used by %s" % reserved[code]))
				continue
			if code in assigned and assigned[code] != item:
				conflicts.append((item, code, "already given to %s" % assigned[code]))
				continue
			seen.add(code)
			assigned[code] = item
			idx += 1
			out.append((code, 1 if code == item else 0, source or "Manual", idx))
		# Guarantee the item's own code is present as the single Primary.
		if item not in seen:
			idx += 1
			out.append((item, 1, "Primary", idx))
		if not any(p for (_, p, _, _) in out):
			for n, (c, p, s, i) in enumerate(out):
				if c == item:
					out[n] = (c, 1, s, i)
					break
		desired[item] = out

	# A "missing" id may actually be an item that RENAMED itself: the id isn't an
	# item name, but one of its file codes currently IS an item name. Flag those so
	# they can be renamed back (they're fixable, not truly absent).
	renamed = {}
	for item in missing:
		for code, _isp, _src in groups[item]:
			if code != item and code in item_names:
				renamed[item] = code  # file id -> the existing (renamed) item
				break

	stats = {
		"file_items": len(order),
		"existing": len(import_items),
		"missing": missing,
		"renamed": renamed,
		"conflicts": conflicts,
		"rows": sum(len(v) for v in desired.values()),
	}
	return import_items, desired, stats


def build_reconcile_plan(groups, order):
	"""Match each file group to a real item by ANY of its codes (so a manually
	renamed item is still found), then plan to rename that item to its PRIMARY code
	and set the full code set. Returns (code_tables, renames, stats).

	Rules per group:
	  - Its codes currently belong to 0 items  -> missing (skipped).
	  - ... to more than 1 item                -> ambiguous / needs manual merge (skipped).
	  - ... to exactly 1 item                  -> reconcile: rename that item to the
	                                              primary code and set its code table.
	"""
	item_names = set(frappe.db.sql_list("select name from `tabItem`"))
	alias_map = {}
	for code, parent in frappe.db.sql(
		"select code, parent from `tabItem Code Entry` where parenttype = 'Item'"
	):
		alias_map.setdefault(code, parent)

	def owner_of(code):
		if code in item_names:
			return code
		return alias_map.get(code)

	code_tables = {}       # final_name -> [(code, is_primary, source, idx)]
	renames = []           # (current_name, final_name)
	missing, ambiguous, conflicts = [], [], []
	renamed_info = {}      # file_id -> current name (informational)
	claimed_target = {}    # final_name -> file_id
	assigned_code = {}     # code -> final_name

	for gid in order:
		# de-dup codes, keep order, find the primary (is_primary=1) else the id
		codes, seen, primary = [], set(), None
		for code, is_primary, source in groups[gid]:
			if code in seen:
				continue
			seen.add(code)
			codes.append((code, is_primary, source))
			if is_primary and primary is None:
				primary = code
		if primary is None:
			primary = gid if gid in seen else (codes[0][0] if codes else gid)

		owners = {owner_of(c) for c, _, _ in codes} - {None}
		if not owners:
			missing.append(gid)
			continue
		if len(owners) > 1:
			ambiguous.append((gid, sorted(owners)))
			continue

		current = next(iter(owners))
		final = primary
		if current != gid:
			renamed_info[gid] = current

		# The final (primary) name must not belong to a DIFFERENT existing item.
		fo = owner_of(final)
		if fo and fo != current:
			conflicts.append((gid, final, "target name already used by %s" % fo))
			continue
		if final in claimed_target and claimed_target[final] != gid:
			conflicts.append((gid, final, "two file rows target item %s" % final))
			continue

		out, local_seen, idx = [], set(), 0
		for code, _is_primary, source in codes:
			o = owner_of(code)
			if o and o != current:
				conflicts.append((gid, code, "code used by %s" % o))
				continue
			if code in assigned_code and assigned_code[code] != final:
				conflicts.append((gid, code, "code already given to %s" % assigned_code[code]))
				continue
			if code in local_seen:
				continue
			local_seen.add(code)
			assigned_code[code] = final
			idx += 1
			out.append((code, 1 if code == final else 0, source or "Manual", idx))
		if final not in local_seen:
			idx += 1
			out.append((final, 1, "Primary", idx))
			assigned_code[final] = final
		if not any(p for (_, p, _, _) in out):
			for n, (c, p, s, i) in enumerate(out):
				if c == final:
					out[n] = (c, 1, s, i)
					break

		code_tables[final] = out
		claimed_target[final] = gid
		if current != final:
			renames.append((current, final))

	stats = {
		"file_items": len(order),
		"existing": len(code_tables),
		"missing": missing,
		"renamed": renamed_info,
		"renames": renames,
		"ambiguous": ambiguous,
		"conflicts": conflicts,
		"rows": sum(len(v) for v in code_tables.values()),
	}
	return code_tables, renames, stats


def apply_reconcile(code_tables, renames, progress=None):
	"""Rename items to their primary code, then rebuild every target's code table."""
	# 1) renames (updates all links + runs the app's after_rename)
	for n, (frm, to) in enumerate(renames, start=1):
		try:
			frappe.rename_doc("Item", frm, to, force=True, show_alert=False, rebuild_search=False)
		except Exception:
			frappe.log_error(title="reconcile rename failed: %s -> %s" % (frm, to))
		if n % 50 == 0:
			frappe.db.commit()
			if progress:
				progress("renamed %d/%d items" % (n, len(renames)))
	frappe.db.commit()

	# 2) replace the code tables (keyed by the final/primary name)
	targets = list(code_tables.keys())
	for i in range(0, len(targets), DELETE_BATCH):
		chunk = targets[i:i + DELETE_BATCH]
		ph = ", ".join(["%s"] * len(chunk))
		frappe.db.sql(
			"delete from `tabItem Code Entry` where parenttype = 'Item' and parent in (%s)" % ph,
			tuple(chunk),
		)
		frappe.db.commit()

	now = now_datetime()
	fields = [
		"name", "creation", "modified", "modified_by", "owner", "docstatus",
		"idx", "parent", "parentfield", "parenttype", "code", "is_primary", "source", "changed_on",
	]
	values = []
	for target, out in code_tables.items():
		for code, is_primary, source, idx in out:
			values.append((
				frappe.generate_hash(length=10), now, now, "Administrator", "Administrator", 0,
				idx, target, PARENT_FIELD, "Item", code, is_primary, source, now,
			))
	for i in range(0, len(values), INSERT_CHUNK):
		frappe.db.bulk_insert(CHILD_DT, fields, values[i:i + INSERT_CHUNK])
		frappe.db.commit()
		if progress:
			progress("inserted %d/%d rows" % (min(i + INSERT_CHUNK, len(values)), len(values)))
	return len(values)


def apply_plan(import_items, desired, progress=None):
	"""Replace each item's code table with the desired rows, in batches."""
	now = now_datetime()
	for i in range(0, len(import_items), DELETE_BATCH):
		chunk = import_items[i:i + DELETE_BATCH]
		ph = ", ".join(["%s"] * len(chunk))
		frappe.db.sql(
			"delete from `tabItem Code Entry` where parenttype = 'Item' and parent in (%s)" % ph,
			tuple(chunk),
		)
		frappe.db.commit()
		if progress:
			progress("deleted %d/%d items" % (min(i + DELETE_BATCH, len(import_items)), len(import_items)))

	fields = [
		"name", "creation", "modified", "modified_by", "owner", "docstatus",
		"idx", "parent", "parentfield", "parenttype", "code", "is_primary", "source", "changed_on",
	]
	values = []
	for item, out in desired.items():
		for code, is_primary, source, idx in out:
			values.append((
				frappe.generate_hash(length=10), now, now, "Administrator", "Administrator", 0,
				idx, item, PARENT_FIELD, "Item", code, is_primary, source, now,
			))
	for i in range(0, len(values), INSERT_CHUNK):
		frappe.db.bulk_insert(CHILD_DT, fields, values[i:i + INSERT_CHUNK])
		frappe.db.commit()
		if progress:
			progress("inserted %d/%d rows" % (min(i + INSERT_CHUNK, len(values)), len(values)))
	return len(values)


def build_report_csv(stats):
	"""CSV bytes listing EVERYTHING that was not updated: missing items + conflicts."""
	import csv
	import io

	buf = io.StringIO()
	w = csv.writer(buf)
	w.writerow(["type", "item_id", "code", "reason"])
	renamed = stats.get("renamed", {})
	for m in stats["missing"]:
		if m in renamed:
			w.writerow(["renamed", m, renamed[m],
			            "found under code '%s' — will be renamed to its primary" % renamed[m]])
		else:
			w.writerow(["missing", m, "", "no item with any of these codes on this site"])
	for gid, items in stats.get("ambiguous", []):
		w.writerow(["ambiguous", gid, ", ".join(items), "codes belong to more than one item — manual merge needed"])
	for item, code, reason in stats["conflicts"]:
		w.writerow(["conflict", item, code, reason])
	return buf.getvalue().encode("utf-8")


def _save_report(stats):
	"""Attach the not-updated CSV to the Item Code Import single doc, return its URL."""
	# Drop the previous report so attachments don't pile up.
	for old in frappe.get_all(
		"File",
		filters={
			"attached_to_doctype": "Item Code Import",
			"attached_to_name": "Item Code Import",
			"file_name": ("like", "item_code_not_updated%"),
		},
		pluck="name",
	):
		frappe.delete_doc("File", old, force=1, ignore_permissions=True)
	f = frappe.get_doc(
		{
			"doctype": "File",
			"file_name": "item_code_not_updated.csv",
			"is_private": 1,
			"content": build_report_csv(stats),
			"attached_to_doctype": "Item Code Import",
			"attached_to_name": "Item Code Import",
		}
	).insert(ignore_permissions=True)
	return f.file_url


def skipped_count(stats):
	"""Everything that will NOT be updated: truly-missing + ambiguous + conflicts."""
	renamed = stats.get("renamed", {})
	truly_missing = len([m for m in stats["missing"] if m not in renamed])
	return truly_missing + len(stats.get("ambiguous", [])) + len(stats["conflicts"])


def summarize(stats, applied_rows=None):
	"""Human-readable summary string for the DocType / CLI."""
	renamed = stats.get("renamed", {})
	renames = stats.get("renames", [])
	ambiguous = stats.get("ambiguous", [])
	# In replace-mode the renamed ids sit INSIDE `missing`; in reconcile-mode they
	# don't. Counting "missing but not renamed" is correct for both.
	truly_missing = len([m for m in stats["missing"] if m not in renamed])
	lines = [
		"Items in file:     %d" % stats["file_items"],
		"Existing items:    %d  (updated)" % stats["existing"],
		"Missing items:     %d  (skipped)" % truly_missing,
		"Likely renamed:    %d  (present under another code)" % len(renamed),
		"Renames to apply:  %d  (item -> its primary code)" % len(renames),
		"Ambiguous:         %d  (codes span >1 item — needs manual merge)" % len(ambiguous),
		"Code conflicts:    %d  (skipped)" % len(stats["conflicts"]),
		"Rows planned:      %d" % stats["rows"],
	]
	if renames:
		lines.append("\nRenames (first 20):")
		lines += ["  - %s  ->  %s" % (a, b) for a, b in renames[:20]]
	elif renamed:
		lines.append("\nLikely renamed (first 20):")
		lines += ["  - %s  ->  now '%s'" % (k, v) for k, v in list(renamed.items())[:20]]
	if ambiguous:
		lines.append("\nAmbiguous (first 20):")
		lines += ["  - %s : codes span %s" % (gid, ", ".join(items)) for gid, items in ambiguous[:20]]
	if applied_rows is not None:
		lines.append("Rows written:   %d" % applied_rows)
	only_missing = [m for m in stats["missing"] if m not in renamed]
	if only_missing:
		lines.append("\nMissing (first 20):")
		lines += ["  - %s" % m for m in only_missing[:20]]
	if stats["conflicts"]:
		lines.append("\nConflicts (first 20):")
		lines += ["  - %s : %s (%s)" % c for c in stats["conflicts"][:20]]
	return "\n".join(lines)


def run_from_path(path, commit=False, progress=None):
	"""End-to-end from a file path. Returns (stats, applied_rows)."""
	groups, order = parse_workbook(path)
	import_items, desired, stats = build_plan(groups, order)
	applied = None
	if commit:
		applied = apply_plan(import_items, desired, progress=progress)
		frappe.db.commit()
	return stats, applied


# --------------------------------------------------------------------------- #
# Desk (Item Code Import DocType) — upload a file and run from the browser.
# --------------------------------------------------------------------------- #
def _file_path(file_url):
	if not file_url:
		frappe.throw(frappe._("Attach a spreadsheet first."))
	return frappe.get_doc("File", {"file_url": file_url}).get_full_path()


def _plan(groups, order, mode):
	"""(stats,) for the chosen mode. reconcile matches by any code + renames."""
	if mode == "reconcile":
		_, _, stats = build_reconcile_plan(groups, order)
	else:
		_, _, stats = build_plan(groups, order)
	return stats


@frappe.whitelist()
def preview(file_url: str, sheet: str | None = None, mode: str | None = None) -> str:
	"""Dry run: parse + plan, attach a full CSV of what won't be updated, return a
	summary. Writes no item data."""
	groups, order = parse_workbook(_file_path(file_url), sheet)
	stats = _plan(groups, order, mode or "replace")
	summary = summarize(stats)
	report_url = _save_report(stats)
	skipped = skipped_count(stats)
	summary += "\n\nFull list of the %d skipped rows is in the attached CSV (Report field / Attachments)." % skipped
	frappe.db.set_single_value("Item Code Import", "result", summary)
	frappe.db.set_single_value("Item Code Import", "report", report_url)
	frappe.db.set_single_value(
		"Item Code Import", "status",
		"Dry run: %d to update, %d skipped" % (stats["existing"], skipped),
	)
	return summary


@frappe.whitelist()
def start(file_url: str, sheet: str | None = None, mode: str | None = None) -> str:
	"""Queue the real import as a background job (handles large files)."""
	_file_path(file_url)  # validate now
	frappe.db.set_single_value("Item Code Import", "status", "Queued...")
	frappe.enqueue(
		"azzir_fleet.item_code_import.background_run",
		queue="long",
		timeout=3600,
		file_url=file_url,
		sheet=sheet,
		mode=mode or "replace",
		user=frappe.session.user,
	)
	return "queued"


def background_run(file_url, sheet=None, mode=None, user=None):
	"""Worker job: apply the import and record the result on the Single doctype."""
	def _progress(msg):
		frappe.db.set_single_value("Item Code Import", "status", "Running: " + msg)
		frappe.db.commit()

	try:
		frappe.db.set_single_value("Item Code Import", "status", "Running: reading file...")
		frappe.db.commit()
		groups, order = parse_workbook(_file_path(file_url), sheet)
		if (mode or "replace") == "reconcile":
			code_tables, renames, stats = build_reconcile_plan(groups, order)
			applied = apply_reconcile(code_tables, renames, progress=_progress)
		else:
			import_items, desired, stats = build_plan(groups, order)
			applied = apply_plan(import_items, desired, progress=_progress)
		frappe.db.commit()
		summary = summarize(stats, applied)
		report_url = _save_report(stats)
		summary += "\n\nFull list of skipped rows is in the attached CSV."
		frappe.db.set_single_value("Item Code Import", "report", report_url)
		frappe.db.set_single_value("Item Code Import", "result", summary)
		frappe.db.set_single_value("Item Code Import", "status", "Completed: %d rows written" % applied)
		frappe.db.commit()
		if user:
			frappe.publish_realtime(
				"msgprint",
				{"message": frappe._("Item code import complete: %d rows.") % applied, "title": "Done"},
				user=user,
			)
	except Exception:
		frappe.db.rollback()
		frappe.db.set_single_value("Item Code Import", "status", "Failed — see Error Log")
		frappe.db.commit()
		frappe.log_error(title="Item Code Import failed")
		if user:
			frappe.publish_realtime(
				"msgprint",
				{"message": frappe._("Item code import failed — check Error Log."), "title": "Failed"},
				user=user,
			)
